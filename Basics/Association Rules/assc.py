from openpyxl import load_workbook
import numpy as np

wb = load_workbook(filename="association_rules.xlsx")
sheetnm = wb.sheetnames

ws1 = wb[sheetnm[0]]
ws2 = wb[sheetnm[1]]

data1 = ws1['D3'].value
data2 = ws1['D4'].value
data3 = ws2['D3'].value
data4 = ws2['D5'].value

data3_mx = np.matrix(data3)
data4_mx = np.matrix(data4)

print(data1)
print(data2)
print(data3_mx)
print(data4_mx)

data_fus = zip(data3_mx, data4_mx)
print(data_fus)
print(list(data_fus))